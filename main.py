import os
from datetime import datetime, timedelta, timezone
from typing import List, Optional

from fastapi import FastAPI, HTTPException, Depends, status, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
import jwt
from jwt import InvalidTokenError
from passlib.context import CryptContext
from pydantic import BaseModel, EmailStr, Field

from database import db, create_document
from schemas import User as UserSchema, Cashflow as CashflowSchema

# ----------------------
# App & CORS
# ----------------------
app = FastAPI(title="Cashflow Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ----------------------
# Auth / Security Setup
# ----------------------
SECRET_KEY = os.getenv("SECRET_KEY", "super-secret-key-change-me")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")


class Token(BaseModel):
    access_token: str
    token_type: str = "bearer"


class RegisterRequest(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: Optional[str] = None


class CashflowCreate(BaseModel):
    amount: float
    reason: str


class CashflowOut(BaseModel):
    id: str
    user_id: str
    amount: float
    reason: str
    created_at: datetime
    updated_at: datetime


# ----------------------
# Utility Functions
# ----------------------

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt


def get_user_by_email(email: str) -> Optional[dict]:
    if db is None:
        raise HTTPException(status_code=500, detail="Database not available")
    user = db["user"].find_one({"email": email})
    return user


def get_user_by_id(user_id: str) -> Optional[dict]:
    from bson import ObjectId

    if db is None:
        raise HTTPException(status_code=500, detail="Database not available")
    try:
        user = db["user"].find_one({"_id": ObjectId(user_id)})
        return user
    except Exception:
        return None


async def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except InvalidTokenError:
        raise credentials_exception
    user = get_user_by_id(user_id)
    if user is None:
        raise credentials_exception
    return user


# ----------------------
# Excel writer
# ----------------------
EXCEL_FILE_PATH = os.getenv("EXCEL_FILE_PATH", "data.xlsx")


def append_to_excel(amount: float, reason: str, user_email: str, created_at: datetime):
    """
    Append a row to the Excel file.
    If the file doesn't exist, create it with headers.
    """
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as e:
        # If openpyxl isn't available, skip silently but allow DB write
        # Raise a clear error so the user can install it if they want Excel support
        raise HTTPException(status_code=500, detail=f"Excel support not available: {e}")

    if not os.path.exists(EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cashflow"
        ws.append(["timestamp", "user", "amount", "reason"])  # headers
        wb.save(EXCEL_FILE_PATH)

    wb = load_workbook(EXCEL_FILE_PATH)
    ws = wb.active
    ws.append([created_at.astimezone(timezone.utc).isoformat(), user_email, amount, reason])
    wb.save(EXCEL_FILE_PATH)


# ----------------------
# Auth Endpoints
# ----------------------
@app.post("/auth/register", response_model=Token)
def register(payload: RegisterRequest):
    existing = get_user_by_email(payload.email)
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    user_doc = UserSchema(email=payload.email, password_hash=get_password_hash(payload.password), name=payload.name)
    user_id = create_document("user", user_doc)

    token = create_access_token({"sub": user_id})
    return Token(access_token=token)


@app.post("/auth/login", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends()):
    user = get_user_by_email(form_data.username)
    if not user:
        raise HTTPException(status_code=400, detail="Incorrect email or password")
    if not verify_password(form_data.password, user.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="Incorrect email or password")

    token = create_access_token({"sub": str(user["_id"])})
    return Token(access_token=token)


@app.get("/me")
def me(current_user: dict = Depends(get_current_user)):
    return {
        "id": str(current_user["_id"]),
        "email": current_user.get("email"),
        "name": current_user.get("name"),
        "is_active": current_user.get("is_active", True),
    }


# ----------------------
# Cashflow Endpoints
# ----------------------
@app.get("/cashflows", response_model=List[CashflowOut])
def list_cashflows(limit: int = 50, current_user: dict = Depends(get_current_user)):
    docs = db["cashflow"].find({"user_id": str(current_user["_id"])}) \
        .sort("created_at", -1).limit(int(limit))

    items: List[CashflowOut] = []
    for d in docs:
        items.append(CashflowOut(
            id=str(d["_id"]),
            user_id=d["user_id"],
            amount=float(d["amount"]),
            reason=d["reason"],
            created_at=d["created_at"],
            updated_at=d["updated_at"],
        ))
    return items


@app.post("/cashflows", response_model=CashflowOut)
def create_cashflow(payload: CashflowCreate, current_user: dict = Depends(get_current_user)):
    user_id = str(current_user["_id"])

    cf_doc = CashflowSchema(user_id=user_id, amount=payload.amount, reason=payload.reason)
    inserted_id = create_document("cashflow", cf_doc)

    # Retrieve the newly created document to return full fields with timestamps
    new_doc = db["cashflow"].find_one({"_id": __import__("bson").ObjectId(inserted_id)})

    # Try appending to Excel; if it fails due to missing package, we don't want to block DB write
    try:
        append_to_excel(cf_doc.amount, cf_doc.reason, current_user.get("email", ""), new_doc.get("created_at", datetime.now(timezone.utc)))
    except HTTPException:
        # best-effort; keep DB success even if Excel write fails
        pass

    # Notify websocket subscribers
    CashflowConnectionManager.broadcast_to_user(user_id, {
        "id": str(new_doc["_id"]),
        "user_id": user_id,
        "amount": float(new_doc["amount"]),
        "reason": new_doc["reason"],
        "created_at": new_doc["created_at"].isoformat() if hasattr(new_doc["created_at"], 'isoformat') else str(new_doc["created_at"]),
        "updated_at": new_doc["updated_at"].isoformat() if hasattr(new_doc["updated_at"], 'isoformat') else str(new_doc["updated_at"]),
    })

    return CashflowOut(
        id=str(new_doc["_id"]),
        user_id=user_id,
        amount=float(new_doc["amount"]),
        reason=new_doc["reason"],
        created_at=new_doc["created_at"],
        updated_at=new_doc["updated_at"],
    )


# ----------------------
# WebSocket for real-time updates
# ----------------------
class CashflowConnectionManager:
    connections_by_user: dict = {}

    @classmethod
    def connect(cls, user_id: str, websocket: WebSocket):
        if user_id not in cls.connections_by_user:
            cls.connections_by_user[user_id] = set()
        cls.connections_by_user[user_id].add(websocket)

    @classmethod
    def disconnect(cls, user_id: str, websocket: WebSocket):
        try:
            cls.connections_by_user.get(user_id, set()).discard(websocket)
        except Exception:
            pass

    @classmethod
    def broadcast_to_user(cls, user_id: str, message: dict):
        for ws in list(cls.connections_by_user.get(user_id, set())):
            try:
                import anyio
                anyio.from_thread.run(ws.send_json, message)
            except Exception:
                # best-effort
                pass


@app.websocket("/ws/cashflows")
async def cashflow_ws(websocket: WebSocket):
    # Expect token as query param: /ws/cashflows?token=...
    token = websocket.query_params.get("token")
    if not token:
        await websocket.close(code=1008)
        return

    # Validate token and extract user
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if not user_id:
            await websocket.close(code=1008)
            return
    except InvalidTokenError:
        await websocket.close(code=1008)
        return

    await websocket.accept()
    CashflowConnectionManager.connect(user_id, websocket)

    try:
        while True:
            # Keep connection alive; we don't expect messages from client
            await websocket.receive_text()
    except WebSocketDisconnect:
        CashflowConnectionManager.disconnect(user_id, websocket)


# ----------------------
# Health & Test
# ----------------------
@app.get("/")
def read_root():
    return {"message": "Cashflow Backend Running"}


@app.get("/test")
def test_database():
    """Test endpoint to check if database is available and accessible"""
    response = {
        "backend": "✅ Running",
        "database": "❌ Not Available",
        "database_url": None,
        "database_name": None,
        "connection_status": "Not Connected",
        "collections": []
    }

    try:
        if db is not None:
            response["database"] = "✅ Available"
            response["database_url"] = "✅ Configured"
            response["database_name"] = db.name if hasattr(db, 'name') else "✅ Connected"
            response["connection_status"] = "Connected"
            try:
                collections = db.list_collection_names()
                response["collections"] = collections[:10]
                response["database"] = "✅ Connected & Working"
            except Exception as e:
                response["database"] = f"⚠️  Connected but Error: {str(e)[:50]}"
        else:
            response["database"] = "⚠️  Available but not initialized"

    except Exception as e:
        response["database"] = f"❌ Error: {str(e)[:50]}"

    import os as _os
    response["database_url"] = "✅ Set" if _os.getenv("DATABASE_URL") else "❌ Not Set"
    response["database_name"] = "✅ Set" if _os.getenv("DATABASE_NAME") else "❌ Not Set"

    return response


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
